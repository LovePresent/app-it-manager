import io
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.asset import Asset
from app.models.category import AssetCategory


def export_assets_to_excel(db: Session, category_id: int | None = None) -> io.BytesIO:
    query = db.query(Asset)
    if category_id:
        query = query.filter(Asset.category_id == category_id)
    assets = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = [
        "자산태그", "이름", "카테고리", "상태", "시리얼번호",
        "모델", "제조사", "구매일", "구매가", "보증만료일", "비고"
    ]
    ws.append(headers)

    for asset in assets:
        cat = db.query(AssetCategory).filter(AssetCategory.id == asset.category_id).first()
        ws.append([
            asset.asset_tag,
            asset.name,
            cat.name if cat else "",
            asset.status,
            asset.serial_number or "",
            asset.model or "",
            asset.manufacturer or "",
            str(asset.purchase_date) if asset.purchase_date else "",
            asset.purchase_price or "",
            str(asset.warranty_expiry) if asset.warranty_expiry else "",
            asset.notes or "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_import_template() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"

    headers = [
        "이름*", "카테고리슬러그*", "상태", "시리얼번호",
        "모델", "제조사", "구매일(YYYY-MM-DD)", "구매가",
        "보증만료일(YYYY-MM-DD)", "비고"
    ]
    ws.append(headers)
    ws.append(["예시 노트북", "laptop", "in_stock", "SN123456", "ThinkPad X1", "Lenovo", "2026-01-15", "1500000", "2029-01-15", "개발팀용"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_assets_from_excel(db: Session, file_bytes: bytes) -> dict:
    from app.services.asset_service import generate_asset_tag
    from datetime import date

    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            slug = str(row[1]).strip() if row[1] else None
            if not slug:
                errors.append(f"행 {i}: 카테고리 슬러그 누락")
                continue

            cat = db.query(AssetCategory).filter(AssetCategory.slug == slug).first()
            if not cat:
                errors.append(f"행 {i}: 존재하지 않는 카테고리 '{slug}'")
                continue

            asset = Asset(
                asset_tag=generate_asset_tag(db),
                name=name,
                category_id=cat.id,
                status=str(row[2]).strip() if row[2] else "in_stock",
                serial_number=str(row[3]).strip() if row[3] else None,
                model=str(row[4]).strip() if row[4] else None,
                manufacturer=str(row[5]).strip() if row[5] else None,
                purchase_date=date.fromisoformat(str(row[6]).strip()) if row[6] else None,
                purchase_price=float(row[7]) if row[7] else None,
                warranty_expiry=date.fromisoformat(str(row[8]).strip()) if row[8] else None,
                notes=str(row[9]).strip() if row[9] else None,
            )
            db.add(asset)
            db.flush()
            created += 1
        except Exception as e:
            errors.append(f"행 {i}: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors}
